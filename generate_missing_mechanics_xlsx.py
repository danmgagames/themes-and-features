"""
Generate output/missing_mechanics_review.xlsx — list of mechanics found in
Pragmatic Play data but absent from our taxonomy, with Spanish descriptions
for Product team review.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

OUTPUT_PATH = Path(__file__).resolve().parent / 'output' / 'missing_mechanics_review.xlsx'

ROWS = [
    {
        'mechanic': 'Hyperplay',
        'pp_count': 137,
        'description_en': 'Pragmatic Play proprietary mechanic where reels expand during play, increasing the number of paylines or symbols available. Conceptually similar to Megaways.',
        'description_es': 'Mecánica propia de Pragmatic Play donde los rodillos se expanden durante el juego, aumentando el número de líneas de pago o símbolos disponibles. Conceptualmente similar a Megaways.',
        'closest_tag': 'Megaways / Multiple Grids',
        'recommendation': 'Add only if MGA has a non-Megaways equivalent (e.g. reel-expansion games outside the Megaways licence)',
    },
    {
        'mechanic': 'Increasing Wilds',
        'pp_count': 8,
        'description_en': 'Wild symbols that increase in number during a single round or feature, accumulating progressively as the round continues.',
        'description_es': 'Símbolos comodín que aumentan en número durante una ronda o función, acumulándose progresivamente a medida que avanza la ronda.',
        'closest_tag': 'Sticky Wild + Random Wild (combined behaviour)',
        'recommendation': 'Add as a distinct tag in the Wilds category — it describes a recognisable mechanic',
    },
    {
        'mechanic': 'Mystery Expanding Symbol',
        'pp_count': 7,
        'description_en': 'A mystery symbol that, when revealed, expands to cover additional reel positions, transforming into a paying or wild symbol.',
        'description_es': 'Un símbolo misterioso que, al revelarse, se expande para cubrir posiciones adicionales en los rodillos, transformándose en un símbolo de pago o comodín.',
        'closest_tag': 'Mystery Symbols + Expanding Wild',
        'recommendation': 'Add as discrete mechanic if MGA has games that combine mystery reveal + expansion',
    },
    {
        'mechanic': 'Powernudge',
        'pp_count': 6,
        'description_en': 'Pragmatic Play branded nudge mechanic where reels shift after a near-miss or specific trigger, attempting to complete winning combinations.',
        'description_es': 'Mecánica de empuje (nudge) de marca Pragmatic Play donde los rodillos se desplazan tras un fallo cercano o un trigger específico, intentando completar combinaciones ganadoras.',
        'closest_tag': 'Nudge & Hold',
        'recommendation': 'Skip — our existing "Nudge & Hold" tag covers this generically',
    },
    {
        'mechanic': 'Super Scatter',
        'pp_count': 2,
        'description_en': 'An enhanced scatter symbol that occupies multiple reel positions or has greater triggering power than a standard scatter symbol.',
        'description_es': 'Un símbolo scatter mejorado que ocupa múltiples posiciones en los rodillos o tiene mayor poder de activación que un símbolo scatter estándar.',
        'closest_tag': 'Scatter Pays',
        'recommendation': 'Optional — low PP frequency (2 games); only add if MGA has clear examples',
    },
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Missing Mechanics'

    headers = [
        'Mechanic (EN)',
        'PP usage count',
        'Description (EN)',
        'Descripción (ES)',
        'Closest existing tag',
        'Recommendation',
        'In MGA games? (Y/N/Notes)',
    ]
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F4F4F')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    for r in ROWS:
        ws.append([
            r['mechanic'],
            r['pp_count'],
            r['description_en'],
            r['description_es'],
            r['closest_tag'],
            r['recommendation'],
            '',
        ])

    widths = [25, 12, 60, 60, 30, 50, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.row_dimensions[1].height = 30
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 90

    ws.freeze_panes = 'A2'

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f'Wrote {OUTPUT_PATH} ({len(ROWS)} mechanics)')


if __name__ == '__main__':
    main()
