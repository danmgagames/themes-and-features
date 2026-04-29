"""
Generate output/themes_features_by_market.xlsx — one sheet per market with
columns: GameName, Category, Themes, Features, Description.

Joins:
  AM_Masterlist (market sheets, localised names) →
  market_names.xlsx (commercial name → tablename → base_key) →
  output/games_enriched.csv (themes, features, description by base_key)
"""

import csv
import json
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from rapidfuzz import fuzz, process

PROJECT_ROOT = Path(__file__).resolve().parent
AM_PATH = PROJECT_ROOT / 'config' / 'AM_Masterlist.xlsx'
MN_PATH = PROJECT_ROOT / 'config' / 'market_names.xlsx'
ENRICHED_PATH = PROJECT_ROOT / 'output' / 'games_enriched.csv'
OUTPUT_PATH = PROJECT_ROOT / 'output' / 'themes_features_by_market.xlsx'
TAXONOMY_PATH = PROJECT_ROOT / 'config' / 'seo_taxonomy.json'
CORRECTIONS_PATH = PROJECT_ROOT / 'output' / 'celebrity_corrections.csv'

SUFFIX_PATTERN = re.compile(r'(NoIp|Es|Pt|It|Co|Nl|Ca|Se)$', re.IGNORECASE)
FUZZY_THRESHOLD = 88
XMARKET_FUZZY_THRESHOLD = 92
CELEBRITIES_UMBRELLA = 'Celebrities'
_CONJUNCTION_RE = re.compile(r'\s*(?:&|\band\b|\by\b|\be\b)\s*')


def norm(s) -> str:
    if s is None:
        return ''
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(s.split())


def norm_match(s) -> str:
    """Like norm() but also collapses & / and / y / e (Spanish, Portuguese,
    Italian conjunctions) so 'Andy & Lucas', 'Andy y Lucas', 'Andy and Lucas'
    all converge for substring matching."""
    n = norm(s)
    return _CONJUNCTION_RE.sub(' & ', n).strip()


def load_taxonomy_themes() -> set[str]:
    """Canonical theme tags = umbrella keys + every sub-tag they list."""
    with open(TAXONOMY_PATH, 'r', encoding='utf-8') as f:
        tx = json.load(f)
    out: set[str] = set()
    themes = tx.get('themes') or {}
    for umbrella, sub_list in themes.items():
        out.add(umbrella)
        for tag in sub_list:
            out.add(tag)
    return out


def collect_celebrity_pool(enriched: dict, taxonomy_themes: set[str]) -> set[str]:
    """Theme tags that co-occur with `Celebrities` somewhere AND are not in
    the canonical taxonomy. These are the celebrity-name tags."""
    pool: set[str] = set()
    for r in enriched.values():
        themes = [t for t in (r.get('themes') or '').split('|') if t]
        if CELEBRITIES_UMBRELLA not in themes:
            continue
        for t in themes:
            if t == CELEBRITIES_UMBRELLA or t in taxonomy_themes:
                continue
            pool.add(t)
    return pool


def validate_celebrities(
    themes_str: str,
    game_name: str,
    celebrity_pool: set[str],
    taxonomy_themes: set[str],
) -> tuple[str, list[dict]]:
    """Strict full-name policy:
      - Drop any celebrity-name tag whose normalized form is not a substring
        of the normalized GameName.
      - Add any celebrity tag from the global pool that IS a substring of
        GameName but is missing from the row's themes.
      - If no celebrity-name tag survives, drop the `Celebrities` umbrella.
      - If a swap-in introduces the first celebrity-name, restore the umbrella.
    """
    if not themes_str:
        return themes_str, []

    themes = [t for t in themes_str.split('|') if t]
    gn_norm = norm_match(game_name)
    log: list[dict] = []

    current_cel_names = [t for t in themes
                         if t != CELEBRITIES_UMBRELLA and t not in taxonomy_themes]

    for cel in current_cel_names:
        if norm_match(cel) not in gn_norm:
            themes.remove(cel)
            log.append({'action': 'remove', 'tag': cel,
                        'reason': 'celebrity name not in localised GameName'})

    for cel in celebrity_pool:
        if cel in themes:
            continue
        cel_norm = norm_match(cel)
        if cel_norm and cel_norm in gn_norm:
            themes.append(cel)
            log.append({'action': 'add', 'tag': cel,
                        'reason': 'celebrity name found in localised GameName'})

    surviving_names = [t for t in themes
                       if t != CELEBRITIES_UMBRELLA and t not in taxonomy_themes]
    if not surviving_names and CELEBRITIES_UMBRELLA in themes:
        themes.remove(CELEBRITIES_UMBRELLA)
        log.append({'action': 'drop_umbrella', 'tag': CELEBRITIES_UMBRELLA,
                    'reason': 'no celebrity-name tag survived row validation'})
    elif surviving_names and CELEBRITIES_UMBRELLA not in themes:
        themes.append(CELEBRITIES_UMBRELLA)
        log.append({'action': 'add', 'tag': CELEBRITIES_UMBRELLA,
                    'reason': 'umbrella restored alongside swap-in celebrity'})

    return '|'.join(themes), log


def load_mn_lookup() -> dict:
    """Build lookup: (market_upper, normalized_commercial_name) → base_key."""
    wb = openpyxl.load_workbook(str(MN_PATH), read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    lookup = {}
    by_market_norms = {}

    for raw in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, raw))
        market = str(d.get('MARKET') or '').upper().strip()
        cname = str(d.get('COMMERCIAL NAME') or '').strip()
        tablename = str(d.get('tablename') or '').strip()
        if not (market and cname and tablename):
            continue
        base_key = SUFFIX_PATTERN.sub('', tablename)
        n = norm(cname)
        lookup[(market, n)] = base_key
        by_market_norms.setdefault(market, []).append((n, base_key))

    wb.close()
    return lookup, by_market_norms


def load_enriched() -> dict:
    """Load enriched CSV, return dict by base_key."""
    by_base = {}
    if not ENRICHED_PATH.exists():
        return by_base
    with open(ENRICHED_PATH, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            bk = r.get('base_key', '').strip()
            if bk:
                by_base[bk] = r
    return by_base


def find_base_key(am_name: str, market: str, mn_lookup: dict, mn_market_norms: dict) -> str | None:
    """Match AM GameName → mn base_key for the given market.

    Tries within-market exact, within-market fuzzy ≥88, then falls back to
    cross-market exact and cross-market fuzzy ≥92. The cross-market fallbacks
    let an AM row inherit a base_key from another market's commercial-name
    listing — useful when a game is sold under the same English title across
    .COM/PORTUGAL/etc but only one market has the entry in market_names.xlsx.
    """
    n = norm(am_name)
    if not n:
        return None
    market_upper = market.upper()

    bk = mn_lookup.get((market_upper, n))
    if bk:
        return bk

    candidates = mn_market_norms.get(market_upper, [])
    if candidates:
        norms_list = [c[0] for c in candidates]
        res = process.extractOne(n, norms_list, scorer=fuzz.token_sort_ratio)
        if res and res[1] >= FUZZY_THRESHOLD:
            return candidates[res[2]][1]

    for (m, nn), bk in mn_lookup.items():
        if nn == n:
            return bk

    all_pairs = []
    for lst in mn_market_norms.values():
        all_pairs.extend(lst)
    if all_pairs:
        names = [p[0] for p in all_pairs]
        res = process.extractOne(n, names, scorer=fuzz.token_sort_ratio)
        if res and res[1] >= XMARKET_FUZZY_THRESHOLD:
            return all_pairs[res[2]][1]

    return None


def main():
    if not AM_PATH.exists():
        raise SystemExit(f"AM_Masterlist not found: {AM_PATH}")
    if not ENRICHED_PATH.exists():
        raise SystemExit(f"Enriched CSV not found: {ENRICHED_PATH}. Run consolidate first.")

    mn_lookup, mn_market_norms = load_mn_lookup()
    enriched = load_enriched()
    print(f"Loaded {len(enriched)} enriched games, {len(mn_lookup)} mn (market, name) lookups")

    taxonomy_themes = load_taxonomy_themes()
    celebrity_pool = collect_celebrity_pool(enriched, taxonomy_themes)
    print(f"Celebrity pool: {len(celebrity_pool)} tags")
    corrections: list[dict] = []

    am_wb = openpyxl.load_workbook(str(AM_PATH), read_only=True, data_only=True)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F4F4F')

    summary = []

    for sheet_name in am_wb.sheetnames:
        am_ws = am_wb[sheet_name]
        rows_iter = am_ws.iter_rows(values_only=True)
        am_headers = next(rows_iter, None)
        if not am_headers:
            continue

        out_ws = out_wb.create_sheet(title=sheet_name[:31])
        out_ws.append(['GameName', 'Category', 'Themes', 'Features', 'Description'])
        for cell in out_ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        market_key = sheet_name.upper().strip()
        n_total = 0
        n_matched = 0

        for raw in rows_iter:
            if not raw:
                continue
            game_name = raw[0]
            if not game_name:
                continue
            n_total += 1
            category = raw[1] if len(raw) > 1 else ''

            bk = find_base_key(str(game_name), market_key, mn_lookup, mn_market_norms)
            themes = ''
            features = ''
            description = ''
            if bk and bk in enriched:
                themes = enriched[bk].get('themes', '')
                features = enriched[bk].get('features', '')
                description = enriched[bk].get('description', '')
                n_matched += 1
                themes, log_entries = validate_celebrities(
                    themes, str(game_name), celebrity_pool, taxonomy_themes)
                for entry in log_entries:
                    corrections.append({
                        'sheet': sheet_name,
                        'base_key': bk,
                        'game_name': str(game_name).strip(),
                        **entry,
                    })

            out_ws.append([
                str(game_name).strip(),
                str(category or ''),
                themes,
                features,
                description,
            ])

        out_ws.column_dimensions['A'].width = 40
        out_ws.column_dimensions['B'].width = 14
        out_ws.column_dimensions['C'].width = 50
        out_ws.column_dimensions['D'].width = 60
        out_ws.column_dimensions['E'].width = 80
        out_ws.freeze_panes = 'A2'

        summary.append((sheet_name, n_total, n_matched))

    am_wb.close()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    target = OUTPUT_PATH
    try:
        out_wb.save(str(target))
    except PermissionError:
        # File likely open in Excel — write to a sibling path so the run still
        # produces an artefact the user can compare against.
        target = OUTPUT_PATH.with_name(OUTPUT_PATH.stem + '.LATEST.xlsx')
        out_wb.save(str(target))
        print()
        print(f'NOTE: Could not overwrite {OUTPUT_PATH.name} (likely open in Excel).')
        print(f'      Wrote to {target.name} instead — close Excel and rename.')

    print()
    print(f"Wrote {OUTPUT_PATH}")
    print(f"{'Market':<14} {'Rows':>6} {'Enriched':>9} {'Coverage':>9}")
    for name, total, matched in summary:
        pct = (matched / total * 100) if total else 0
        print(f"{name:<14} {total:>6} {matched:>9} {pct:>8.1f}%")

    fields = ['sheet', 'base_key', 'game_name', 'action', 'tag', 'reason']
    CORRECTIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CORRECTIONS_PATH, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for row in sorted(corrections,
                          key=lambda r: (r['sheet'], r['base_key'], r['action'], r['tag'])):
            w.writerow(row)

    by_action: dict[str, int] = {}
    for c in corrections:
        by_action[c['action']] = by_action.get(c['action'], 0) + 1
    print()
    print(f"Wrote {CORRECTIONS_PATH.relative_to(PROJECT_ROOT)} ({len(corrections)} rows)")
    print(f"  removals: {by_action.get('remove', 0)} | "
          f"additions: {by_action.get('add', 0)} | "
          f"umbrella drops: {by_action.get('drop_umbrella', 0)}")


if __name__ == '__main__':
    main()
