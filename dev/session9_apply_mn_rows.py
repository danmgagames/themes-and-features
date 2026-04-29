"""
Append 20 alias / new-base_key rows to config/market_names.xlsx so the
remaining Bucket A AM Masterlist rows resolve to a base_key.

Each row is one of:
- ALIAS to existing classified base_key (e.g., 'Hawaii 5-0' → S3HawaiiCountersGlobal)
- NEW base_key that gets classified this session (e.g., Nacho Vidal,
  Ruleta Magic Red)
- Cross-language localised variant pointing at a canonical SPAIN base_key
  via a per-market suffix (Dream 3 Team .COM/Pt/Co, NL Far West Mania,
  NL Take the Money)

Idempotent — won't add a duplicate if (MARKET, COMMERCIAL NAME, tablename)
already exists.
"""
import shutil
import sys
from pathlib import Path
from datetime import datetime

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
MN_PATH = PROJECT_ROOT / 'config' / 'market_names.xlsx'

# (name, category, COMMERCIAL_NAME, MARKET, Gameid, tablename, Enum, Deactivated, rate)
ROWS = [
    # Phase A — alias to classified existing base_keys
    ('TroyaMariaLaPiedraAlias', 'SLOTS3', 'Maria La Piedra En Troya', 'SPAIN', '', 'S3TroyaCountersGlobal', '', 'NO', 'Premium'),
    ('HawaiiAlias', 'SLOTS3', 'Hawaii 5-0', 'SPAIN', '', 'S3HawaiiCountersGlobal', '', 'NO', 'Premium'),
    ('ElCartelNavidadAlias', 'SLOTS3', 'El Cartel Plus Navidad', 'SPAIN', '', 'S3ElCartelNavidadCountersGlobal', '', 'NO', 'Premium'),
    ('MinaDeOroHalloweenAlias', 'SLOTS3', 'La Mina De Oro Plus Halloween', 'SPAIN', '', 'S3MinaDeOroHalloweenCountersGlobal', '', 'NO', 'Premium'),
    ('MinaDeOroNavidadAlias', 'SLOTS3', 'La Mina De Oro Plus Navidad', 'SPAIN', '', 'S3MinaDeOroNavidadCountersGlobal', '', 'NO', 'Premium'),
    ('RFReinaCleopatraAlias', 'SLOTS3', 'Rf Reinas De Africa Cleopatra', 'SPAIN', '', 'S3RFReinaCleopatraCountersGlobal', '', 'NO', 'Premium'),
    ('GrandCroupierMariaLapiedraAlias', 'ROULETTE', 'Ruleta Grand Croupier Sc', 'SPAIN', '', 'RouletteGeneralCountersGrandCroupierOnlyMariaLapiedra', '', 'NO', 'Premium'),
    ('PopeyePTAlias', 'SLOTS3', 'Popeye Caça Tesouros', 'PORTUGAL', '', 'S3PopeyeCountersGlobal', '', 'NO', 'Premium'),
    ('PoliDiazBoxingChampionES', 'MEGAWAYS', 'Poli Diaz Boxing Champion', 'SPAIN', '', 'MegawaysGeneralCountersBoxingPoliDiaz', '', 'NO', 'Premium'),
    ('GrandCroupierMariaLapiedraNoIp', 'ROULETTE', 'Ruleta Grand Croupier Sc', '.COM', '', 'RouletteGeneralCountersGrandCroupierOnlyMariaLapiedraNoIp', '', 'NO', 'Premium'),
    ('FarWestManiaNl', 'MEGAWAYS', 'Ayla de Zwart Far West Mania Megaways', 'NETHERLANDS', '', 'MegawaysGeneralCountersFarWestManiaNl', '', 'NO', 'Premium'),
    ('TakeTheMoneyMegawaysNlAlias', 'MEGAWAYS', 'Neem het Geld Megaways', 'NETHERLANDS', '', 'MegawaysGeneralCountersTakeTheMoneyNl', '', 'NO', 'Premium'),

    # Phase B — new base_keys (will be classified in this session)
    ('AramisFusterLaBrujaAlias', 'SLOTS3', 'Aramis Fuster La Bruja', 'SPAIN', '', 'S3AramisFusterCountersGlobal', '', 'NO', 'Premium'),
    ('Dream3TeamAlias', 'SLOTS5', 'Dream 3 Team', 'SPAIN', '', 'Slots5GeneralCountersDream3Team', '', 'NO', 'Premium'),
    ('Dream3TeamPtAlias', 'SLOTS5', 'Dream 3 Team', 'PORTUGAL', '', 'Slots5GeneralCountersDream3TeamPt', '', 'NO', 'Premium'),
    ('Dream3TeamNoIp', 'SLOTS5', 'Dream3Team', '.COM', '', 'Slots5GeneralCountersDream3TeamNoIp', '', 'NO', 'Premium'),
    ('Dream3TeamCo', 'SLOTS5', 'Dream 3 Team', 'COLOMBIA', '', 'Slots5GeneralCountersDream3TeamCo', '', 'NO', 'Premium'),
    ('NachoVidalMegaways', 'MEGAWAYS', 'Nacho Vidal', 'SPAIN', '', 'MegawaysGeneralCountersNachoVidal', '', 'NO', 'Premium'),
    ('RuletaMagicRedES', 'ROULETTE', 'Ruleta Magic Red', 'SPAIN', '', 'RouletteGeneralCountersMagicRed', '', 'NO', 'Premium'),
    ('RuletaMagicRedNoIp', 'ROULETTE', 'Ruleta Magic Red', '.COM', '', 'RouletteGeneralCountersMagicRedNoIp', '', 'NO', 'Premium'),
]


def main():
    if not MN_PATH.exists():
        sys.exit(f'market_names.xlsx not found: {MN_PATH}')

    backup = MN_PATH.with_suffix(f'.bak.session9-{datetime.now().strftime("%Y%m%d-%H%M%S")}.xlsx')
    shutil.copy2(MN_PATH, backup)
    print(f'Backed up to {backup.name}')

    wb = openpyxl.load_workbook(str(MN_PATH))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f'Headers: {headers}')
    expected = ['name', 'category', 'COMMERCIAL NAME', 'MARKET', 'Gameid', 'tablename', 'Enum', 'Deactivated', 'rate']
    if headers != expected:
        sys.exit(f'Unexpected header order. Got: {headers}\nExpected: {expected}')

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        existing.add((str(d.get('MARKET') or '').strip().upper(),
                      str(d.get('COMMERCIAL NAME') or '').strip().lower(),
                      str(d.get('tablename') or '').strip()))

    added = 0; skipped = 0
    for r in ROWS:
        key = (r[3].upper().strip(), r[2].lower().strip(), r[5].strip())
        if key in existing:
            print(f'  SKIP (already exists): {r[3]:<12} "{r[2]}" -> {r[5]}')
            skipped += 1
            continue
        ws.append(list(r))
        existing.add(key)
        added += 1
        print(f'  ADD: {r[3]:<12} "{r[2]:<40}" -> tablename={r[5]}')

    wb.save(str(MN_PATH))
    print()
    print(f'Done. Added {added}, skipped {skipped} of {len(ROWS)} proposed rows.')
    print(f'Backup: {backup.name}')


if __name__ == '__main__':
    main()
