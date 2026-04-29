"""
Phase 1b: PDF → raw JSON extraction for Session 6.

Walks X:\\DivisionOnline\\FinalesProducto\\<MARKET>\\<CATEGORY>\\<GameName>\\,
locates the per-game "Descripcion del juego" PDF using the priority rules
documented in dev/ref/session6-plan.md, extracts text via pypdf (with
pdfplumber fallback), matches the folder name against market_names.xlsx
to obtain a canonical base_key, and writes one JSON per game family to
the output directory.

When the same game family has PDFs in multiple market subdirs, the best
candidate is chosen by language priority (ENG > ES > BZL) and tie-broken
by family-match confidence.
"""

import json
import logging
import re
import unicodedata
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz, process
from tqdm import tqdm

from agents.localisation_resolver import (
    build_game_families,
    detect_celebrity_ips,
    load_market_db,
    match_extract_to_family,
)

SUFFIX_PATTERN = re.compile(r'(NoIp|Es|Pt|It|Co|Nl|Ca|Se)$', re.IGNORECASE)
PER_MARKET_FUZZY_THRESHOLD = 88

# Tokens that AM Masterlist GameNames append/prepend but market_names omits
# (or vice versa). Used by the loose-match fallback. Edge tokens only — never
# strip from the middle of a name. RF prefix is the "Rey de la Fiesta" tag.
GENERIC_EDGE_TOKENS = {'bingo', 'megaways', 'plus', 'deluxe', 'rf'}


def _norm(s) -> str:
    if s is None:
        return ''
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(s.split())


def _norm_loose(n: str) -> str:
    """Strip leading/trailing generic edge tokens AND inner whitespace from
    an already-_norm'd string. Used for the bridge fallback that catches AM
    GameNames like "Carnaval Bingo" matching market_names "Carnaval", or
    "Mr Magnifico" matching "MrMagnifico". Returns '' if everything was
    generic (don't match on empty)."""
    if not n:
        return ''
    toks = n.split()
    while toks and toks[0] in GENERIC_EDGE_TOKENS:
        toks = toks[1:]
    while toks and toks[-1] in GENERIC_EDGE_TOKENS:
        toks = toks[:-1]
    return ''.join(toks)


def _build_commercial_lookup(market_names_path: Path) -> tuple[dict, dict]:
    """
    Build a per-market commercial-name → base_key lookup.

    Returns:
      exact: dict[(market_upper, normalized_cname)] -> base_key
      by_market: dict[market_upper] -> list[(normalized_cname, base_key)]

    This sidesteps the build_game_families quirk where the SPAIN commercial
    name can be masked by the .COM English commercial name when families
    are aggregated.
    """
    wb = openpyxl.load_workbook(str(market_names_path), read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    exact: dict = {}
    by_market: dict = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, raw))
        market = str(d.get('MARKET') or '').upper().strip()
        cname = str(d.get('COMMERCIAL NAME') or '').strip()
        tablename = str(d.get('tablename') or '').strip()
        if not (market and cname and tablename):
            continue
        base_key = SUFFIX_PATTERN.sub('', tablename)
        n = _norm(cname)
        exact[(market, n)] = base_key
        by_market.setdefault(market, []).append((n, base_key))
    wb.close()
    return exact, by_market


def _resolve_base_key_per_market(
    folder_name: str,
    market_label: str,
    exact: dict,
    by_market: dict,
) -> tuple[str | None, float, str]:
    """
    Resolve folder_name → base_key using per-market commercial-name lookup.

    Tries:
      1. Exact (market, normalized) match
      2. Fuzzy match within the market sublist (token_sort_ratio >= 88)
      3. Cross-market exact (any market) — last resort

    Returns (base_key | None, confidence, method).
    """
    n = _norm(folder_name)
    if not n:
        return None, 0.0, 'none'

    market_upper = (market_label or '').upper()

    # 1. Exact within target market
    bk = exact.get((market_upper, n))
    if bk:
        return bk, 1.0, 'mn_exact'

    # 2. Fuzzy within target market
    candidates = by_market.get(market_upper, [])
    if candidates:
        names = [c[0] for c in candidates]
        res = process.extractOne(n, names, scorer=fuzz.token_sort_ratio)
        if res and res[1] >= PER_MARKET_FUZZY_THRESHOLD:
            return candidates[res[2]][1], res[1] / 100.0, 'mn_fuzzy'

    # 3. Cross-market exact (any market) — handles non-Spain canonical folders
    for (m, nn), bk in exact.items():
        if nn == n:
            return bk, 0.95, 'mn_exact_xmarket'

    # 4. Cross-market fuzzy — only if very strong (>=92)
    all_pairs = []
    for m, lst in by_market.items():
        all_pairs.extend(lst)
    if all_pairs:
        names = [c[0] for c in all_pairs]
        res = process.extractOne(n, names, scorer=fuzz.token_sort_ratio)
        if res and res[1] >= 92:
            return all_pairs[res[2]][1], res[1] / 100.0, 'mn_fuzzy_xmarket'

    # 5. Within-market loose: strip generic edge tokens + inner whitespace.
    #    Catches "Carnaval Bingo" (AM) ↔ "Carnaval" (MN) and "Mr Magnifico"
    #    (AM) ↔ "MrMagnifico" (MN). Only fires when stricter steps failed.
    n_loose = _norm_loose(n)
    if n_loose:
        for cn_norm, bk in candidates:
            if _norm_loose(cn_norm) == n_loose:
                return bk, 0.85, 'mn_loose'

        # 6. Cross-market loose
        for cn_norm, bk in all_pairs:
            if _norm_loose(cn_norm) == n_loose:
                return bk, 0.80, 'mn_loose_xmarket'

    return None, 0.0, 'none'

logger = logging.getLogger(__name__)

# Map FinalesProducto market subdir name → AM-style market label
MARKET_FOLDERS = {
    'ESPAÑA': 'SPAIN',
    'PORTUGAL': 'PORTUGAL',
    'ITALIA': 'ITALY',
    'COLOMBIA': 'COLOMBIA',
    'PAISES BAJOS': 'NETHERLANDS',
    'TODOS MERCADOS': '.COM',
    'OTROS MERCADOS': '.COM',
}

# Categories under each market
CATEGORY_FOLDERS = {'BINGO', 'MEGAWAYS', 'SLOTS3', 'SLOTS5', 'RULETA'}

# Language priority — lower wins
LANG_PRIORITY = {'EN': 0, 'ES': 1, 'BZL': 2, 'UNK': 3}

# Filename patterns for language detection
LANG_PATTERNS = [
    (re.compile(r'_ENG?(_|\.|\b)', re.IGNORECASE), 'EN'),
    (re.compile(r'_ESP?(_|\.|\b)', re.IGNORECASE), 'ES'),
    (re.compile(r'_BZL?(_|\.|\b)', re.IGNORECASE), 'BZL'),
    (re.compile(r'\bENG\b', re.IGNORECASE), 'EN'),
    (re.compile(r'\bESP\b', re.IGNORECASE), 'ES'),
]

# Subdir name patterns (case-insensitive)
DESC_DIR_RE = re.compile(r'descripcion\s*del?\s*juego', re.IGNORECASE)
DESC_FILE_RE = re.compile(r'descripcion[\s_-]*juego', re.IGNORECASE)


def _detect_language(filename: str) -> str:
    for pat, lang in LANG_PATTERNS:
        if pat.search(filename):
            return lang
    return 'UNK'


def _find_description_pdf(game_folder: Path) -> tuple[Path | None, str]:
    """
    Locate the best description PDF for a game folder.

    Priority:
      1. Marketing Assets/06. Descripcion del juego/*.pdf  (SLOTS5 convention)
      2. Gamesheets/*Descripcion del juego*.pdf or descripcion_juego_*.pdf  (BINGO)
      3. Any subfolder matching 'descripcion del juego' / 'descripcion juego'
         containing a relevant PDF
      4. Any PDF anywhere under the game folder whose filename matches
         'descripcion'+'juego'

    Within each rule, language preference is ENG > ES > BZL > UNK.
    Returns (path or None, language).
    """
    candidates: list[tuple[Path, str]] = []

    # Rule 1: Marketing Assets/06. Descripcion del juego/
    marketing = game_folder / 'Marketing Assets'
    if marketing.is_dir():
        for sub in marketing.iterdir():
            if sub.is_dir() and DESC_DIR_RE.search(sub.name):
                for pdf in sub.glob('*.pdf'):
                    candidates.append((pdf, _detect_language(pdf.name)))
    if candidates:
        candidates.sort(key=lambda c: LANG_PRIORITY.get(c[1], 9))
        return candidates[0]

    # Rule 2: Gamesheets/*Descripcion juego*.pdf
    gamesheets = game_folder / 'Gamesheets'
    if gamesheets.is_dir():
        for pdf in gamesheets.glob('*.pdf'):
            if DESC_FILE_RE.search(pdf.name):
                candidates.append((pdf, _detect_language(pdf.name)))
    if candidates:
        candidates.sort(key=lambda c: LANG_PRIORITY.get(c[1], 9))
        return candidates[0]

    # Rule 3: any subfolder named like "Descripcion del juego"
    for sub in game_folder.rglob('*'):
        if not sub.is_dir():
            continue
        if DESC_DIR_RE.search(sub.name):
            for pdf in sub.glob('*.pdf'):
                candidates.append((pdf, _detect_language(pdf.name)))
    if candidates:
        candidates.sort(key=lambda c: LANG_PRIORITY.get(c[1], 9))
        return candidates[0]

    # Rule 4: any PDF in the tree whose filename matches descripcion+juego
    for pdf in game_folder.rglob('*.pdf'):
        if DESC_FILE_RE.search(pdf.name):
            candidates.append((pdf, _detect_language(pdf.name)))
    if candidates:
        candidates.sort(key=lambda c: LANG_PRIORITY.get(c[1], 9))
        return candidates[0]

    return None, 'UNK'


def _extract_pdf_text(pdf_path: Path) -> str:
    """Extract text from PDF using pypdf; fall back to pdfplumber if blank."""
    text = ''
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        parts = []
        for page in reader.pages:
            try:
                parts.append(page.extract_text() or '')
            except Exception:
                continue
        text = '\n'.join(parts).strip()
    except Exception as e:
        logger.warning(f"pypdf failed on {pdf_path}: {e}")

    if text:
        return text

    try:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf:
            parts = [(p.extract_text() or '') for p in pdf.pages]
        text = '\n'.join(parts).strip()
    except Exception as e:
        logger.warning(f"pdfplumber failed on {pdf_path}: {e}")

    return text


def _walk_finales(root: Path) -> list[dict]:
    """Yield game-folder records: {market, market_label, category, folder}."""
    records = []
    for market_dir in sorted(root.iterdir()) if root.is_dir() else []:
        if not market_dir.is_dir():
            continue
        market_label = MARKET_FOLDERS.get(market_dir.name)
        if not market_label:
            continue
        for category_dir in sorted(market_dir.iterdir()):
            if not category_dir.is_dir():
                continue
            cat_upper = category_dir.name.upper()
            if cat_upper not in CATEGORY_FOLDERS:
                continue
            for game_folder in sorted(category_dir.iterdir()):
                if not game_folder.is_dir():
                    continue
                records.append({
                    'market': market_dir.name,
                    'market_label': market_label,
                    'category': cat_upper,
                    'folder': game_folder,
                })
    return records


def _safe_filename(s: str) -> str:
    return re.sub(r'[^\w\-]', '_', s)


def extract_all_pdfs(
    finales_root: Path,
    market_names_path: Path,
    output_dir: Path,
    survey_csv_path: Path | None = None,
    enriched_skip_set: set[str] | None = None,
    overwrite: bool = False,
) -> dict:
    """
    Walk finales_root, find description PDFs, match to game families, write
    one JSON per family to output_dir.

    Returns a stats dict.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    enriched_skip_set = enriched_skip_set or set()

    # Load families for context (es_commercial_name, markets, celebrities)
    market_rows = load_market_db(market_names_path)
    families = build_game_families(market_rows)
    detect_celebrity_ips(families)
    logger.info(f"Loaded {len(families)} game families")

    # Build a per-market commercial-name lookup for robust folder matching.
    # This sidesteps the families quirk where SPAIN cname can be masked by
    # an earlier .COM English cname.
    cname_exact, cname_by_market = _build_commercial_lookup(market_names_path)
    logger.info(f"Per-market commercial-name lookup: {len(cname_exact)} entries")

    records = _walk_finales(finales_root)
    logger.info(f"Discovered {len(records)} game folders across markets")

    survey_rows: list[dict] = []
    candidates_by_base_key: dict[str, list[dict]] = {}
    unmatched: list[dict] = []

    for rec in tqdm(records, desc="Surveying PDFs"):
        folder = rec['folder']
        pdf_path, language = _find_description_pdf(folder)
        pdf_found = pdf_path is not None

        # Primary match: per-market commercial-name lookup
        clean_name = folder.name.replace('_', ' ').strip()
        base_key, match_conf, match_method = _resolve_base_key_per_market(
            clean_name, rec['market_label'], cname_exact, cname_by_market,
        )

        # Fallback: families-based matcher (catches base_key/name-column patterns
        # the per-market lookup misses)
        if not base_key:
            fake_extract = {
                'folder_game_name': clean_name,
                'folder_category': rec['category'],
                'market_lookup': {},
            }
            fam_match = match_extract_to_family(fake_extract, families)
            if fam_match.get('base_key'):
                base_key = fam_match['base_key']
                match_conf = fam_match.get('match_confidence', 0.0)
                match_method = f"family_{fam_match.get('match_method', '')}"

        survey_rows.append({
            'market': rec['market'],
            'market_label': rec['market_label'],
            'category': rec['category'],
            'game_folder': folder.name,
            'pdf_path': str(pdf_path) if pdf_path else '',
            'language': language if pdf_found else '',
            'bytes': pdf_path.stat().st_size if pdf_path and pdf_path.exists() else 0,
            'matched_base_key': base_key or '',
            'match_method': match_method,
            'match_confidence': match_conf,
        })

        if not pdf_found:
            continue

        if not base_key:
            unmatched.append({
                'folder': folder.name,
                'market': rec['market'],
                'category': rec['category'],
                'pdf_path': str(pdf_path),
            })
            continue

        candidates_by_base_key.setdefault(base_key, []).append({
            'pdf_path': pdf_path,
            'language': language,
            'market': rec['market'],
            'market_label': rec['market_label'],
            'category': rec['category'],
            'folder_name': folder.name,
            'match_confidence': match_conf,
            'match_method': match_method,
        })

    # Write coverage survey
    if survey_csv_path is not None:
        survey_csv_path.parent.mkdir(parents=True, exist_ok=True)
        import csv
        cols = [
            'market', 'market_label', 'category', 'game_folder',
            'pdf_path', 'language', 'bytes',
            'matched_base_key', 'match_method', 'match_confidence',
        ]
        with open(survey_csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows(survey_rows)

    # Pick best candidate per base_key and extract
    stats = {
        'folders_scanned': len(records),
        'pdfs_found': sum(1 for r in survey_rows if r['pdf_path']),
        'unmatched_folders': len(unmatched),
        'unique_families_with_pdf': len(candidates_by_base_key),
        'extracts_written': 0,
        'extracts_skipped_existing': 0,
        'extract_errors': 0,
    }

    for base_key, candidates in tqdm(candidates_by_base_key.items(), desc="Extracting PDFs"):
        # Sort: language priority, then highest match_confidence
        candidates.sort(key=lambda c: (
            LANG_PRIORITY.get(c['language'], 9),
            -c['match_confidence'],
        ))
        best = candidates[0]

        safe = _safe_filename(base_key)
        output_path = output_dir / f"{safe}.json"

        # Idempotent: skip when JSON exists AND base_key already enriched
        if output_path.exists() and base_key in enriched_skip_set and not overwrite:
            stats['extracts_skipped_existing'] += 1
            continue

        try:
            raw_text = _extract_pdf_text(best['pdf_path'])
        except Exception as e:
            logger.error(f"Failed extracting {best['pdf_path']}: {e}")
            stats['extract_errors'] += 1
            continue

        fam = families.get(base_key, {})
        record = {
            'base_key': base_key,
            'folder_game_name': best['folder_name'],
            'folder_category': best['category'],
            'market': best['market'],
            'market_label': best['market_label'],
            'pdf_path': str(best['pdf_path']),
            'pdf_found': True,
            'pdf_source_language': best['language'],
            'raw_text': raw_text,
            'raw_text_chars': len(raw_text),
            'all_candidate_pdfs': [
                {
                    'pdf_path': str(c['pdf_path']),
                    'market': c['market'],
                    'language': c['language'],
                }
                for c in candidates
            ],
            'market_lookup': {
                'base_key': base_key,
                'es_commercial_name': fam.get('es_commercial_name'),
                'category': fam.get('category'),
                'markets': fam.get('markets', []),
                'celebrity_names': fam.get('celebrity_names', []),
                'match_confidence': best['match_confidence'],
                'match_method': best['match_method'],
            },
        }

        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(record, f, indent=2, ensure_ascii=False)
            stats['extracts_written'] += 1
        except Exception as e:
            logger.error(f"Failed writing {output_path}: {e}")
            stats['extract_errors'] += 1

    return stats
