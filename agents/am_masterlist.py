"""
AM_Masterlist enrichment.

Loads the SPAIN sheet of config/AM_Masterlist.xlsx and joins each classified
game to a masterlist row to attach current-live spec data (TIER, Release Date,
RTP, Volatility, Max Multiplier, Pay Lines, Reels, Demo URL).

Also produces a gap report listing AM Spain games with no PPTX coverage.
"""

import logging
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz, process

logger = logging.getLogger(__name__)

AM_COLUMNS = [
    'am_tier', 'am_release_date', 'am_rtp', 'am_volatility',
    'am_max_multiplier', 'am_pay_lines', 'am_reels', 'am_demo_url',
    'am_match_method',
]

EMPTY_AM = {col: '' for col in AM_COLUMNS}
EMPTY_AM['am_match_method'] = 'none'

FUZZY_THRESHOLD = 90


def _norm(s) -> str:
    if s is None:
        return ''
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(s.split())


def _fmt_date(v) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v) if v is not None else ''


def load_am_market(am_path: Path, sheet_name: str) -> list[dict]:
    """Load a market sheet from AM_Masterlist.xlsx into row dicts.

    The first column header has a newline in it ('Updated DD/MM/YYYY\\nGameName')
    so we coerce it to 'GameName' for ergonomic access. All 6 market sheets
    follow this convention.
    """
    if not am_path.exists():
        logger.warning(f"AM_Masterlist not found at {am_path}")
        return []

    wb = openpyxl.load_workbook(str(am_path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        logger.warning(f"AM_Masterlist has no {sheet_name} sheet")
        wb.close()
        return []

    ws = wb[sheet_name]
    raw_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not raw_rows:
        return []

    headers = list(raw_rows[0])
    headers[0] = 'GameName'
    headers = [h.strip() if isinstance(h, str) else h for h in headers]

    rows = []
    for raw in raw_rows[1:]:
        row = dict(zip(headers, raw))
        if not row.get('GameName'):
            continue
        rows.append(row)

    rows_with_norm = []
    for row in rows:
        row['_norm_name'] = _norm(row.get('GameName'))
        row['_norm_demo_link'] = _norm(row.get('Demo Link'))
        rows_with_norm.append(row)

    return rows_with_norm


def load_am_spain(am_path: Path) -> list[dict]:
    return load_am_market(am_path, 'SPAIN')


def _row_to_am_dict(row: dict, method: str) -> dict:
    return {
        'am_tier': row.get('TIER') or '',
        'am_release_date': _fmt_date(row.get('Release Date')),
        'am_rtp': row.get('RTP') or '',
        'am_volatility': row.get('Volatility') or '',
        'am_max_multiplier': row.get('Max Multiplier') or '',
        'am_pay_lines': row.get('Pay Lines') or '',
        'am_reels': row.get('Reels') or '',
        'am_demo_url': row.get('Demo URL') or '',
        'am_match_method': method,
    }


def match_classified_to_am(game: dict, am_rows: list[dict]) -> dict:
    """Link a classified game to its AM Spain row.

    Tries (1) exact normalized match on es_commercial_name, (2) exact match on
    folder_game_name, (3) fuzzy token_sort_ratio >= 90 on commercial name.
    """
    if not am_rows:
        return dict(EMPTY_AM)

    candidates = []
    es_name = game.get('es_commercial_name')
    if es_name:
        candidates.append(_norm(es_name))
    folder_name = game.get('folder_game_name')
    if folder_name:
        candidates.append(_norm(folder_name.replace('_', ' ')))

    for cand in candidates:
        if not cand:
            continue
        for row in am_rows:
            if row['_norm_name'] == cand:
                return _row_to_am_dict(row, 'exact')

    am_norms = [r['_norm_name'] for r in am_rows]
    for cand in candidates:
        if not cand:
            continue
        match = process.extractOne(cand, am_norms, scorer=fuzz.token_sort_ratio)
        if match and match[1] >= FUZZY_THRESHOLD:
            return _row_to_am_dict(am_rows[match[2]], 'fuzzy')

    return dict(EMPTY_AM)


def build_gap_report(am_rows: list[dict], classified_games: list[dict]) -> list[dict]:
    """Return AM Spain rows with no matching classified game."""
    matched_names = set()
    for game in classified_games:
        method = game.get('am_match_method', 'none')
        if method == 'none':
            continue
        es = _norm(game.get('es_commercial_name'))
        folder = _norm((game.get('folder_game_name') or '').replace('_', ' '))
        if es:
            matched_names.add(es)
        if folder:
            matched_names.add(folder)

    am_norms_to_match = list(matched_names)

    gap = []
    for row in am_rows:
        norm_name = row['_norm_name']
        if norm_name in matched_names:
            continue
        if am_norms_to_match:
            res = process.extractOne(norm_name, am_norms_to_match, scorer=fuzz.token_sort_ratio)
            if res and res[1] >= FUZZY_THRESHOLD:
                continue
        gap.append({
            'commercial_name': row.get('GameName') or '',
            'category': row.get('Category') or '',
            'tier': row.get('TIER') or '',
            'release_date': _fmt_date(row.get('Release Date')),
            'volatility': row.get('Volatility') or '',
            'rtp': row.get('RTP') or '',
            'demo_url': row.get('Demo URL') or '',
            'notes': 'No PPTX in current download',
        })

    gap.sort(key=lambda r: (r['category'], r['commercial_name'].lower()))
    return gap
