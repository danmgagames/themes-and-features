"""
Phase 2c: Localisation resolution.

Groups game families from market_names.xlsx, resolves localisation variants,
and detects celebrity/IP attachments in Ca/Se markets.

Deterministic lookup first; Claude API fallback for ambiguous matches only.
"""

import json
import logging
import re
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz

logger = logging.getLogger(__name__)

SUFFIX_PATTERN = re.compile(r'(NoIp|Es|Pt|It|Co|Nl|Ca|Se)$', re.IGNORECASE)


def load_market_db(market_names_path: Path) -> list[dict]:
    """Load market_names.xlsx into a list of row dicts."""
    wb = openpyxl.load_workbook(str(market_names_path), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        rows.append(row_dict)
    wb.close()
    return rows


def build_game_families(market_rows: list[dict], include_deactivated: bool = False) -> dict:
    """
    Group market rows into game families keyed by base_key (stripped tablename).

    Returns dict[base_key] = {
        'base_key': str,
        'es_commercial_name': str or None,
        'category': str,
        'markets': [str],
        'variants': [row_dict],
        'celebrity_names': [str],
        'deactivated': bool,
    }
    """
    families = {}

    for row in market_rows:
        tablename = str(row.get('tablename', '') or '')
        if not tablename:
            continue

        deactivated = str(row.get('Deactivated', '')).strip().upper() == 'YES'
        if deactivated and not include_deactivated:
            continue

        base_key = SUFFIX_PATTERN.sub('', tablename)
        market = str(row.get('MARKET', '') or '').strip()
        commercial_name = str(row.get('COMMERCIAL NAME', '') or '').strip()
        category = str(row.get('category', '') or '').strip()

        if base_key not in families:
            families[base_key] = {
                'base_key': base_key,
                'es_commercial_name': None,
                'category': category,
                'markets': [],
                'variants': [],
                'celebrity_names': [],
                'deactivated': False,
            }

        fam = families[base_key]
        fam['variants'].append(row)

        if market and market not in fam['markets']:
            fam['markets'].append(market)

        # Spain entry is canonical
        if market.upper() in ('SPAIN', '.COM', '') or tablename.lower().endswith('es'):
            if not fam['es_commercial_name'] and commercial_name:
                fam['es_commercial_name'] = commercial_name

        if deactivated:
            fam['deactivated'] = True

    # Fallback: if no Spain entry found, use first available commercial name
    for fam in families.values():
        if not fam['es_commercial_name'] and fam['variants']:
            for v in fam['variants']:
                cn = str(v.get('COMMERCIAL NAME', '') or '').strip()
                if cn:
                    fam['es_commercial_name'] = cn
                    break

    return families


def detect_celebrity_ips(families: dict) -> dict:
    """
    Detect celebrity/IP attachments by comparing Ca/Se variant commercial names
    against the base commercial name.

    Modifies families in-place and returns them.
    """
    for base_key, fam in families.items():
        base_name = (fam['es_commercial_name'] or '').strip().lower()
        if not base_name:
            continue

        for variant in fam['variants']:
            market = str(variant.get('MARKET', '') or '').strip().upper()
            if market not in ('CANADA', 'SWEDEN'):
                continue

            variant_name = str(variant.get('COMMERCIAL NAME', '') or '').strip()
            if not variant_name:
                continue

            # If the variant name is different and longer, the prefix is likely a celebrity
            variant_lower = variant_name.lower()
            if variant_lower != base_name and base_name in variant_lower:
                # Extract the prefix (celebrity name)
                idx = variant_lower.index(base_name)
                celebrity = variant_name[:idx].strip()
                if celebrity and celebrity not in fam['celebrity_names']:
                    fam['celebrity_names'].append(celebrity)
                    logger.info(f"[{base_key}] Celebrity IP detected: {celebrity} (from {market})")

    return families


def match_extract_to_family(
    extract: dict,
    families: dict,
) -> dict:
    """
    Match a raw extract to a game family from market_names.xlsx.

    Returns match result dict with base_key, es_commercial_name, markets,
    celebrity_names, match_confidence, match_method.
    """
    game_name = extract.get('folder_game_name', '').replace('_', ' ').strip()
    category = extract.get('folder_category', '').strip()

    # If the extract already has a market lookup with a base_key, try that first
    market_lookup = extract.get('market_lookup', {})
    existing_base_key = market_lookup.get('base_key')
    if existing_base_key and existing_base_key in families:
        fam = families[existing_base_key]
        return _build_result(fam, market_lookup.get('match_confidence', 1.0), 'extract_lookup')

    # 1. Exact match on base_key
    for bk, fam in families.items():
        if game_name.lower().replace(' ', '') == bk.lower().replace(' ', ''):
            return _build_result(fam, 1.0, 'exact_base_key')

    # 2. Exact match on es_commercial_name
    for bk, fam in families.items():
        es_name = (fam['es_commercial_name'] or '').lower()
        if es_name and es_name == game_name.lower():
            return _build_result(fam, 1.0, 'exact_commercial')

    # 3. Fuzzy match on es_commercial_name
    best_score = 0
    best_fam = None
    for bk, fam in families.items():
        es_name = fam['es_commercial_name'] or ''
        score = fuzz.token_sort_ratio(game_name.lower(), es_name.lower())
        if score > best_score:
            best_score = score
            best_fam = fam

    if best_score >= 75 and best_fam:
        return _build_result(best_fam, best_score / 100.0, 'fuzzy_commercial')

    # 4. Fuzzy match on base_key
    best_score = 0
    best_fam = None
    for bk, fam in families.items():
        score = fuzz.token_sort_ratio(game_name.lower(), bk.lower())
        if score > best_score:
            best_score = score
            best_fam = fam

    if best_score >= 75 and best_fam:
        return _build_result(best_fam, best_score / 100.0, 'fuzzy_base_key')

    # No match
    return {
        'base_key': None,
        'es_commercial_name': None,
        'category': category or extract.get('folder_category'),
        'markets': [],
        'celebrity_names': [],
        'match_confidence': 0.0,
        'match_method': 'none',
    }


def _build_result(fam: dict, confidence: float, method: str) -> dict:
    return {
        'base_key': fam['base_key'],
        'es_commercial_name': fam['es_commercial_name'],
        'category': fam['category'],
        'markets': fam['markets'],
        'celebrity_names': fam['celebrity_names'],
        'match_confidence': round(confidence, 3),
        'match_method': method,
    }
